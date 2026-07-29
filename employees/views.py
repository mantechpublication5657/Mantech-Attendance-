# employees/views.py
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.core.paginator import Paginator
from django.db.models import Q
from decimal import Decimal
from django.conf import settings

from .forms import (
    EmployeeBasicInfoForm,
    EmployeeAdminForm,
    AddEmployeeForm,
)
from .models import EmployeeProfile, AdminControlledData

User = get_user_model()
from datetime import date
from calendar import monthrange
from attendance.models import Attendance


# --------------------------------------------------
# EMPLOYEE DASHBOARD (unchanged)
# --------------------------------------------------
import json
import csv
import io
import calendar as cal
from calendar import monthrange
from datetime import date,timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Avg, Count
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import render

from .models import EmployeeProfile, AdminControlledData
from attendance.models import Attendance
from payroll.models import Payroll



# Readable department name lookup
DEPARTMENT_DISPLAY = {
    'hr': 'Human Resources',
    'finance': 'Finance',
    'it': 'Information Technology',
    'sales': 'Sales',
    'marketing': 'Marketing',
    'operations': 'Operations',
    'admin': 'Administration',
}



@login_required
def employee_dashboard(request):

    from calendar import monthrange
    from datetime import date, timedelta
    from decimal import Decimal
    import calendar as cal
    import json

    from django.db.models import Sum, Avg, Count

    user = request.user
    profile, _ = EmployeeProfile.objects.get_or_create(user=user)
    today = date.today()

    # =========================================================
    # FILTERS
    # =========================================================
    filter_type = request.GET.get('filter_type', 'monthly')

    try:
        filter_month = int(request.GET.get('month', today.month))
    except (ValueError, TypeError):
        filter_month = today.month

    try:
        filter_year = int(request.GET.get('year', today.year))
    except (ValueError, TypeError):
        filter_year = today.year

    filter_month = max(1, min(12, filter_month))

    # =========================================================
    # HELPER — working days in month (exclude Sundays)
    # Used only for attendance % calculation
    # =========================================================
    def working_days_in_month(year, month):
        total = monthrange(year, month)[1]
        sundays = sum(
            1 for d in range(1, total + 1)
            if date(year, month, d).weekday() == 6
        )
        return max(total - sundays, 1)

    # =========================================================
    # HELPER — 2nd Saturday of a month
    # =========================================================
    def get_second_saturday(year, month):
        saturday_count = 0
        for d in range(1, monthrange(year, month)[1] + 1):
            current = date(year, month, d)
            if current.weekday() == 5:
                saturday_count += 1
                if saturday_count == 2:
                    return current
        return None

    # =========================================================
    # HELPER — attendance deduction
    # EXACT mirror of payroll_detail logic:
    #   per_day_salary = basic / total_calendar_days  (NOT working days)
    #   step 1: bulk deduct ALL absent days
    #   step 2: deduct half-day salary per Half Day (skip 2nd Saturday)
    #   step 3: Sunday rule — deduct if both adjacent Sat AND Mon absent
    # =========================================================
    def calc_attendance_deduction(att_qs, per_day_salary, half_day_salary, second_saturday):
        attendance_deduction = Decimal("0.00")

        # 1. Bulk deduct all absent days (including Sundays — matches payroll_detail)
        absent_days = att_qs.filter(status="Absent").count()
        attendance_deduction += per_day_salary * Decimal(absent_days)

        # 2. Half Day → deduct half UNLESS it falls on 2nd Saturday
        for record in att_qs.filter(status="Half Day"):
            if second_saturday and record.date == second_saturday:
                continue  # 2nd Saturday Half Day = full pay, no deduction
            attendance_deduction += half_day_salary

        # 3. Sunday rule — deduct only if BOTH adjacent Sat AND Mon are absent
        for sunday_record in att_qs.filter(date__week_day=1):
            saturday = sunday_record.date - timedelta(days=1)
            monday   = sunday_record.date + timedelta(days=1)
            sat_absent = att_qs.filter(date=saturday, status="Absent").exists()
            mon_absent = att_qs.filter(date=monday,   status="Absent").exists()
            if sat_absent and mon_absent:
                attendance_deduction += per_day_salary

        # 4. 2nd Saturday — absent already caught in bulk step 1
        #    Half Day already skipped in step 2; nothing extra needed here

        return attendance_deduction.quantize(Decimal("0.01"))

    # =========================================================
    # PERIOD — attendance queryset + working days + label
    # =========================================================
    if filter_type == 'yearly':
        period_attendance = Attendance.objects.filter(
            employee=user,
            date__year=filter_year,
        )
        total_working_days = sum(
            working_days_in_month(filter_year, m) for m in range(1, 13)
        )
        period_label = str(filter_year)
    else:
        period_attendance = Attendance.objects.filter(
            employee=user,
            date__year=filter_year,
            date__month=filter_month,
        )
        total_working_days = working_days_in_month(filter_year, filter_month)
        period_label = f"{cal.month_name[filter_month]} {filter_year}"

    # =========================================================
    # ATTENDANCE COUNTS
    # =========================================================
    present_count  = period_attendance.filter(status="Present").count()
    late_count     = period_attendance.filter(status="Late").count()
    leave_count    = period_attendance.filter(status="Leave").count()
    absent_count   = period_attendance.filter(status="Absent").count()
    half_day_count = period_attendance.filter(status="Half Day").count()

    # =========================================================
    # PAYABLE DAYS — Present=1, Late=1, Leave=1, Half Day=0.5
    # =========================================================
    payable_days = (
        Decimal(present_count) +
        Decimal(late_count) +
        Decimal(leave_count) +
        Decimal(half_day_count) * Decimal("0.5")
    )

    # =========================================================
    # ATTENDANCE %
    # =========================================================
    attendance_pct = (
        min(round((float(payable_days) / float(total_working_days)) * 100), 100)
        if total_working_days > 0 else 0
    )

    # =========================================================
    # LEAVE BALANCE
    # =========================================================
    leave_balance = max(12 - leave_count, 0)

    # =========================================================
    # TODAY'S ATTENDANCE
    # =========================================================
    today_attendance = Attendance.objects.filter(
        employee=user, date=today
    ).first()

    # =========================================================
    # MONTHLY CHART DATA
    # =========================================================
    monthly_att_data      = []
    monthly_sal_data      = []
    monthly_half_day_data = []

    for m in range(1, 13):
        m_att = Attendance.objects.filter(
            employee=user,
            date__year=filter_year,
            date__month=m,
        )
        m_present = m_att.filter(status="Present").count()
        m_late    = m_att.filter(status="Late").count()
        m_leave   = m_att.filter(status="Leave").count()
        m_half    = m_att.filter(status="Half Day").count()
        m_working = working_days_in_month(filter_year, m)

        m_payable = (
            Decimal(m_present) +
            Decimal(m_late) +
            Decimal(m_leave) +
            Decimal(m_half) * Decimal("0.5")
        )

        m_pct = (
            min(round((float(m_payable) / float(m_working)) * 100), 100)
            if m_working > 0 else 0
        )

        monthly_att_data.append(m_pct)
        monthly_half_day_data.append(m_half)

        try:
            payroll_row = Payroll.objects.filter(
                employee=user, month=m, year=filter_year
            ).first()
            monthly_sal_data.append(
                round(float(payroll_row.net_salary) / 100000, 2)
                if payroll_row else 0
            )
        except Exception:
            monthly_sal_data.append(0)

    # =========================================================
    # ADMIN DASHBOARD
    # =========================================================
    admin_context = {}

    if user.is_staff:

        all_profiles    = EmployeeProfile.objects.filter(user__is_active=True)
        total_employees = all_profiles.count()

        # Today snapshot
        present_today  = Attendance.objects.filter(date=today, status="Present").count()
        half_day_today = Attendance.objects.filter(date=today, status="Half Day").count()
        absent_today   = Attendance.objects.filter(date=today, status="Absent").count()

        # Payroll for selected period
        if filter_type == 'yearly':
            period_payrolls = Payroll.objects.filter(year=filter_year)
        else:
            period_payrolls = Payroll.objects.filter(
                year=filter_year, month=filter_month
            )

        total_payroll = period_payrolls.aggregate(t=Sum('net_salary'))['t'] or 0
        avg_salary    = period_payrolls.aggregate(a=Avg('net_salary'))['a'] or 0

        # Overall attendance
        if filter_type == 'yearly':
            all_att = Attendance.objects.filter(date__year=filter_year)
            overall_working = total_employees * sum(
                working_days_in_month(filter_year, m) for m in range(1, 13)
            )
        else:
            all_att = Attendance.objects.filter(
                date__year=filter_year, date__month=filter_month
            )
            overall_working = total_employees * working_days_in_month(
                filter_year, filter_month
            )

        all_present = all_att.filter(status="Present").count()
        all_late    = all_att.filter(status="Late").count()
        all_leave   = all_att.filter(status="Leave").count()
        all_half    = all_att.filter(status="Half Day").count()

        overall_payable = (
            Decimal(all_present) +
            Decimal(all_late) +
            Decimal(all_leave) +
            Decimal(all_half) * Decimal("0.5")
        )

        overall_attendance = (
            min(round((float(overall_payable) / float(overall_working)) * 100), 100)
            if overall_working > 0 else 0
        )

        # Department distribution
        dept_qs = (
            all_profiles
            .exclude(department__isnull=True)
            .values('department')
            .annotate(c=Count('id'))
            .order_by('-c')
        )
        dept_labels = [
            DEPARTMENT_DISPLAY.get(d['department'], d['department'])
            for d in dept_qs
        ]
        dept_counts = [d['c'] for d in dept_qs]

        # Payroll trend (last 4 years)
        payroll_years  = []
        payroll_totals = []
        for y in range(today.year - 3, today.year + 1):
            yr_total = (
                Payroll.objects.filter(year=y)
                .aggregate(t=Sum('net_salary'))['t'] or 0
            )
            payroll_years.append(str(y))
            payroll_totals.append(round(float(yr_total) / 100000, 2))

        # -------------------------------------------------
        # RECENT PAYROLL TABLE
        # Deduction logic mirrors payroll_detail exactly:
        #   per_day_salary = basic / total_calendar_days
        # -------------------------------------------------
        recent_payrolls_qs = (
            period_payrolls
            .select_related('employee')
            .order_by('-id')[:10]
        )

        recent_payrolls = []

        for p in recent_payrolls_qs:

            emp_profile = getattr(p.employee, 'employee_profile', None)
            admin_data  = (
                getattr(emp_profile, 'admin_data', None)
                if emp_profile else None
            )

            # ── Use total calendar days (matches payroll_detail) ──
            p_total_days   = monthrange(p.year, p.month)[1]
            per_day_salary = (p.basic_salary / Decimal(p_total_days)).quantize(Decimal("0.01"))
            half_day_sal   = (per_day_salary / Decimal("2")).quantize(Decimal("0.01"))
            second_sat     = get_second_saturday(p.year, p.month)

            att = Attendance.objects.filter(
                employee=p.employee,
                date__year=p.year,
                date__month=p.month,
            )

            # Counts for display columns
            row_present = att.filter(status__in=["Present", "Late", "Leave"]).count()
            row_half    = att.filter(status="Half Day").count()
            row_absent  = att.filter(status="Absent").count()

            # Attendance deduction — exact mirror of payroll_detail
            total_att_deduction = calc_attendance_deduction(
                att, per_day_salary, half_day_sal, second_sat
            )

            # Total deduction = fixed deductions stored on Payroll + attendance deduction
            total_deductions = (
                p.deductions + total_att_deduction
            ).quantize(Decimal("0.01"))

            recent_payrolls.append({
                'full_name':            p.employee.get_full_name() or p.employee.username,
                'emp_id':               emp_profile.emp_id if emp_profile else '—',
                'department':           DEPARTMENT_DISPLAY.get(emp_profile.department, '—') if emp_profile else '—',
                'designation':          admin_data.designation if admin_data else '—',
                'basic_salary':         p.basic_salary,
                'hra':                  p.hra,
                'da':                   p.da,
                'bonus':                p.bonus,
                'deductions':           p.deductions,
                'attendance_deduction': total_att_deduction,
                'half_days':            row_half,
                'total_deductions':     total_deductions,
                'net_salary':           p.net_salary,
                'month_year':           f"{cal.month_name[p.month]} {p.year}",
                'unpaid_days':          row_absent,
                'present_days':         row_present,
            })

        admin_context = {
            'total_employees':    total_employees,
            'present_today':      present_today,
            'half_day_today':     half_day_today,
            'absent_today':       absent_today,
            'total_payroll':      f"{total_payroll:,.0f}",
            'avg_salary':         f"{avg_salary:,.0f}",
            'overall_attendance': overall_attendance,
            'recent_payrolls':    recent_payrolls,
            'dept_labels':        json.dumps(dept_labels),
            'dept_counts':        json.dumps(dept_counts),
            'payroll_years':      json.dumps(payroll_years),
            'payroll_totals':     json.dumps(payroll_totals),
        }

    # =========================================================
    # FILTER DROPDOWNS
    # =========================================================
    months_list = [cal.month_name[m] for m in range(1, 13)]
    years_list  = list(range(today.year - 3, today.year + 1))

    # =========================================================
    # CONTEXT
    # =========================================================
    context = {
        "employee":              profile,
        "attendance_pct":        attendance_pct,
        "leave_balance":         leave_balance,
        "present_count":         present_count,
        "late_count":            late_count,
        "leave_count":           leave_count,
        "absent_count":          absent_count,
        "half_day_count":        half_day_count,
        "payable_days":          payable_days,
        "today_attendance":      today_attendance,
        "filter_type":           filter_type,
        "filter_month":          filter_month,
        "filter_year":           filter_year,
        "filter_period":         f"{filter_year}-{filter_month:02d}",
        "period_label":          period_label,
        "months_list":           months_list,
        "current_month":         filter_month,
        "current_month_name":    cal.month_name[filter_month],
        "years_list":            years_list,
        "current_year":          filter_year,
        "monthly_att_data":      json.dumps(monthly_att_data),
        "monthly_sal_data":      json.dumps(monthly_sal_data),
        "monthly_half_day_data": json.dumps(monthly_half_day_data),
        **admin_context,
    }

    return render(request, "employees/employee_dashboard.html", context)

# ─── Download view ─────────────────────────────────────────────────────────────

@login_required
def payroll_report_download(request):
    if not request.user.is_staff:
        return HttpResponseForbidden("Admin only")

    fmt         = request.GET.get('format', 'csv')
    filter_type = request.GET.get('filter_type', 'monthly')
    today       = date.today()

    try:
        month = int(request.GET.get('month', today.month))
        year  = int(request.GET.get('year', today.year))
    except (ValueError, TypeError):
        month, year = today.month, today.year

    if filter_type == 'yearly':
        payrolls   = Payroll.objects.filter(year=year).select_related('employee')
        period_str = str(year)
    else:
        payrolls   = Payroll.objects.filter(year=year, month=month).select_related('employee')
        period_str = f"{cal.month_name[month]}-{year}"

    headers = [
        'Emp ID', 'Employee', 'Department', 'Designation',
        'Basic Salary', 'HRA','Earned HRA', 'SA','Earned SA', 'Bonus',
        'Manual Deductions', 'Absence Deduction',
        'Half Days', 'Half Day Deduction',
        'Sunday Cut', 'Total Deductions',
        'Net Salary', 'Period',
    ]

    # =========================================================
    # HELPER — 2nd Saturday of a month
    # =========================================================
    def get_second_saturday(year, month):
        saturday_count = 0
        for d in range(1, cal.monthrange(year, month)[1] + 1):
            current = date(year, month, d)
            if current.weekday() == 5:
                saturday_count += 1
                if saturday_count == 2:
                    return current
        return None

    def build_rows():
        rows = []

        for p in payrolls:

            emp_profile = getattr(p.employee, 'employee_profile', None)
            admin_data  = getattr(emp_profile, 'admin_data', None) if emp_profile else None

            # ── Use total calendar days — matches payroll_detail ──
            total_month_days = cal.monthrange(p.year, p.month)[1]
            per_day_salary   = (p.basic_salary / Decimal(total_month_days)).quantize(Decimal("0.01"))
            half_day_salary  = (per_day_salary / Decimal("2")).quantize(Decimal("0.01"))
            second_saturday  = get_second_saturday(p.year, p.month)

            attendances = Attendance.objects.filter(
                employee=p.employee,
                date__year=p.year,
                date__month=p.month,
            )

            # ── Step 1: Absent deduction — bulk all absent days ──
            # Leave is PAID so excluded; matches payroll_detail step 1
            absent_days       = attendances.filter(status="Absent").count()
            absence_deduction = (
                per_day_salary * Decimal(absent_days)
            ).quantize(Decimal("0.01"))

            # ── Step 2: Half Day — per-record, skip 2nd Saturday ──
            # matches payroll_detail step 2
            half_day_deduction      = Decimal("0.00")
            half_day_deduction_count = 0
            for record in attendances.filter(status="Half Day"):
                if second_saturday and record.date == second_saturday:
                    continue  # 2nd Saturday Half Day = full pay
                half_day_deduction       += half_day_salary
                half_day_deduction_count += 1

            half_day_deduction = half_day_deduction.quantize(Decimal("0.01"))

            # ── Step 3: Sunday rule ──
            # Deduct only if BOTH adjacent Sat AND Mon are Absent
            # (Leave on either side does NOT trigger — matches payroll_detail)
            sunday_cut = Decimal("0.00")
            for sunday_rec in attendances.filter(date__week_day=1):
                sat = sunday_rec.date - timedelta(days=1)
                mon = sunday_rec.date + timedelta(days=1)
                sat_absent = attendances.filter(date=sat, status="Absent").exists()
                mon_absent = attendances.filter(date=mon, status="Absent").exists()
                if sat_absent and mon_absent:
                    sunday_cut += per_day_salary

            sunday_cut = sunday_cut.quantize(Decimal("0.01"))

            # ── Step 4: 2nd Saturday ──
            # Absent → already in bulk step 1
            # Half Day → already skipped in step 2
            # Present/Late/Leave → full pay, nothing extra

            # ── Total deductions ──
            total_deductions = (
                p.deductions
                + absence_deduction
                + half_day_deduction
                + sunday_cut
            ).quantize(Decimal("0.01"))

            rows.append([
                emp_profile.emp_id if emp_profile else '—',
                p.employee.get_full_name() or p.employee.username,
                DEPARTMENT_DISPLAY.get(emp_profile.department, '—') if emp_profile else '—',
                admin_data.designation if admin_data else '—',
                float(p.basic_salary),
                float(p.hra),
                float(p.earned_hra),
                float(p.da),
                float(p.earned_da),
                float(p.bonus),
                float(p.deductions),
                float(absence_deduction),
                half_day_deduction_count,
                float(half_day_deduction),
                float(sunday_cut),
                float(total_deductions),
                float(p.net_salary),
                f"{cal.month_name[p.month]} {p.year}",
            ])

        return rows

    # ── CSV ──────────────────────────────────────────────────────────
    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = (
            f'attachment; filename="payroll_{period_str}.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(build_rows())
        return response

    # ── Excel ────────────────────────────────────────────────────────
    if fmt == 'xlsx':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            return HttpResponse("Run: pip install openpyxl", status=500)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Payroll {period_str}"

        hdr_fill      = PatternFill("solid", fgColor="1E3A8A")
        hdr_font      = Font(bold=True, color="FFFFFF")
        half_day_fill = PatternFill("solid", fgColor="B45309")

        half_day_col_indices = {
            headers.index('Half Days') + 1,
            headers.index('Half Day Deduction') + 1,
        }

        for col, h in enumerate(headers, 1):
            cell           = ws.cell(row=1, column=col, value=h)
            cell.font      = hdr_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill      = half_day_fill if col in half_day_col_indices else hdr_fill

        for ri, row in enumerate(build_rows(), 2):
            for ci, val in enumerate(row, 1):
                cell           = ws.cell(row=ri, column=ci, value=val)
                cell.alignment = Alignment(horizontal='center')
                if ci in half_day_col_indices:
                    cell.fill = PatternFill("solid", fgColor="FEF3C7")
                    cell.font = Font(
                        color="92400E",
                        bold=(ci == headers.index('Half Days') + 1)
                    )

        # Auto column width
        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        response = HttpResponse(
            buf,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = (
            f'attachment; filename="payroll_{period_str}.xlsx"'
        )
        return response

    return HttpResponse("Invalid format", status=400)

# Utility: check if user is admin/staff
def is_admin(user):
    return user.is_staff or user.is_superuser


# --------------------------------------------------
# ADD EMPLOYEE (Admin only)
# --------------------------------------------------
@user_passes_test(is_admin)
def add_employee(request):
    if request.method == "POST":
        form = AddEmployeeForm(request.POST, request.FILES)
        if form.is_valid():
            # Create user
            user = User.objects.create_user(
                username=form.cleaned_data["username"],
                email=form.cleaned_data["email"],
                first_name=form.cleaned_data["first_name"],
                last_name=form.cleaned_data["last_name"],
                password=form.cleaned_data["password1"],
            )

            # Create profile
            profile = EmployeeProfile.objects.create(
                user=user,
                department=form.cleaned_data["department"],
                phone=form.cleaned_data.get("phone"),
                address=form.cleaned_data.get("address"),
                gender=form.cleaned_data.get("gender"),
                date_of_birth=form.cleaned_data.get("date_of_birth"),
                emergency_contact_phone=form.cleaned_data.get("emergency_contact_phone"),
                profile_picture=form.cleaned_data.get("profile_picture"),
            )

            # Create admin-controlled data
            AdminControlledData.objects.create(
                profile=profile,
                joining_date=form.cleaned_data["joining_date"],
                designation=form.cleaned_data.get("designation"),
                role=form.cleaned_data.get("role"),
            )

            messages.success(request, "Employee added successfully.")
            return redirect("employee_list")
    else:
        form = AddEmployeeForm()
    return render(request, "employees/add_employee.html", {"form": form})

# --------------------------------------------------
# EDIT ADD EMPLOYEE (Admin only)
# --------------------------------------------------
@user_passes_test(is_admin)
def edit_add_employee(request, user_id):
    """
    Admin can edit both basic employee info and admin-controlled data.
    """
    user = get_object_or_404(User, id=user_id)
    profile, _ = EmployeeProfile.objects.get_or_create(user=user)
    admin_data, _ = AdminControlledData.objects.get_or_create(profile=profile)

    if request.method == "POST":
        basic_form = EmployeeBasicInfoForm(request.POST, request.FILES, instance=profile)
        admin_form = EmployeeAdminForm(request.POST, instance=admin_data)

        if basic_form.is_valid() and admin_form.is_valid():
            basic_form.save()
            admin_form.save()
            messages.success(request, "Employee details updated successfully.")
            return redirect("employee_detail", user_id=user.id)
    else:
        basic_form = EmployeeBasicInfoForm(instance=profile)
        admin_form = EmployeeAdminForm(instance=admin_data)

    return render(request, "employees/edit_add_employee.html", {
        "profile": profile,
        "basic_form": basic_form,
        "admin_form": admin_form,
        "is_admin": True,
    })

# --------------------------------------------------
# EDIT BASIC INFO (Employee self-edit)
# --------------------------------------------------
@login_required
def edit_employee_basic(request):
    profile = request.user.employee_profile
    if request.method == "POST":
        form = EmployeeBasicInfoForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, "Your profile has been updated.")
            return redirect("employee_detail", user_id=profile.user.id)
    else:
        form = EmployeeBasicInfoForm(instance=profile)
    return render(request, "employees/edit_employee_basic.html", {"form": form})


# --------------------------------------------------
# EMPLOYEE DETAIL VIEW
# --------------------------------------------------
from datetime import date
import calendar
@login_required
def employee_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile, created = EmployeeProfile.objects.get_or_create(user=user)

    today = date.today()
    month = today.month
    year = today.year

    # =========================================
    # PAYROLL (optional — may not exist yet)
    # =========================================
    payroll = Payroll.objects.filter(employee=user, month=month, year=year).first()

    present_days = absent_days = late_days = leave_days = half_day_days = 0
    total_month_days = total_sundays = 0
    second_saturday = None
    per_day_salary = half_day_salary = payable_days = cut_days = Decimal("0.00")
    attendance_deduction = total_deduction = gross_salary = net_salary = Decimal("0.00")

    if payroll:
        payroll.calculate_salary()
        payroll.refresh_from_db()

        # =========================================
        # ATTENDANCE RECORDS FOR THIS MONTH
        # =========================================
        attendances = Attendance.objects.filter(
            employee=user,
            date__year=year,
            date__month=month
        )

        present_days  = attendances.filter(status="Present").count()
        absent_days   = attendances.filter(status="Absent").count()
        late_days     = attendances.filter(status="Late").count()
        leave_days    = attendances.filter(status="Leave").count()
        half_day_days = attendances.filter(status="Half Day").count()

        # =========================================
        # CALENDAR — SUNDAYS & 2ND SATURDAY
        # =========================================
        total_month_days = calendar.monthrange(year, month)[1]
        saturday_count = 0

        for day in range(1, total_month_days + 1):
            current_date = date(year, month, day)
            if current_date.weekday() == 6:
                total_sundays += 1
            if current_date.weekday() == 5:
                saturday_count += 1
                if saturday_count == 2:
                    second_saturday = current_date

        # =========================================
        # PER DAY / HALF DAY SALARY
        # =========================================
        per_day_salary  = (payroll.basic_salary / Decimal(total_month_days)).quantize(Decimal("0.01"))
        half_day_salary = (per_day_salary / Decimal("2")).quantize(Decimal("0.01"))

        # =========================================
        # PAYABLE DAYS
        # =========================================
        payable_days = (
            Decimal(present_days) +
            Decimal(late_days) +
            Decimal(leave_days) +
            Decimal("0.5") * Decimal(half_day_days)
        )

        # =========================================
        # ATTENDANCE DEDUCTION
        # =========================================
        attendance_deduction = Decimal("0.00")

        attendance_deduction += per_day_salary * Decimal(absent_days)

        for record in attendances.filter(status="Half Day"):
            if second_saturday and record.date == second_saturday:
                continue
            attendance_deduction += half_day_salary

        for sunday_record in attendances.filter(date__week_day=1):
            saturday = sunday_record.date - timedelta(days=1)
            monday   = sunday_record.date + timedelta(days=1)
            sat_absent = attendances.filter(date=saturday, status="Absent").exists()
            mon_absent = attendances.filter(date=monday,   status="Absent").exists()
            if sat_absent and mon_absent:
                attendance_deduction += per_day_salary
            else:
                payable_days += Decimal("1")

        # =========================================
        # CUT DAYS
        # =========================================
        half_day_deduction_count = Decimal("0")
        for record in attendances.filter(status="Half Day"):
            if second_saturday and record.date == second_saturday:
                continue
            half_day_deduction_count += Decimal("0.5")

        cut_days = Decimal(absent_days) + half_day_deduction_count

        # =========================================
        # TOTALS
        # =========================================
        total_deduction = (payroll.deductions + attendance_deduction).quantize(Decimal("0.01"))

        gross_salary = (
            payroll.basic_salary +
            payroll.hra +
            payroll.da +
            payroll.bonus
        )

        net_salary = payroll.net_salary

    context = {
        "employee": profile,
        'avatar_list': settings.AVATAR_LIST,
        "month": month,
        "year": year,
        "payroll": payroll,

        # Attendance
        "present_days":   present_days,
        "absent_days":    absent_days,
        "late_days":      late_days,
        "leave_days":     leave_days,
        "half_day_days":  half_day_days,

        # Calendar
        "total_month_days": total_month_days,
        "total_sundays":    total_sundays,
        "second_saturday":  second_saturday,

        # Salary breakdown
        "per_day_salary":   per_day_salary,
        "half_day_salary":  half_day_salary,
        "payable_days":     payable_days,
        "cut_days":         cut_days,

        # Deductions
        "attendance_deduction": attendance_deduction,
        "total_deduction":      total_deduction,

        # Totals
        "gross_salary": gross_salary,
        "net_salary":   net_salary,
    }

    return render(request, "employees/employee_detail.html", context)


# --------------------------------------------------
# EMPLOYEE LIST (Admin only)
# --------------------------------------------------
@user_passes_test(is_admin)
def employee_list(request):
    query = request.GET.get("q", "")
    employees_qs = EmployeeProfile.objects.select_related("user", "admin_data").all()

    if query:
        employees_qs = employees_qs.filter(
            Q(user__username__icontains=query) |
            Q(emp_id__icontains=query)
        )

    paginator = Paginator(employees_qs, 10)
    page_number = request.GET.get("page")
    employees = paginator.get_page(page_number)

    return render(request, "employees/employee_list.html", {"employees": employees})


# --------------------------------------------------
# ADMIN CONTROLLED DATA DETAIL (Admin only)
# --------------------------------------------------
@user_passes_test(is_admin)
def admin_data_detail(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile = get_object_or_404(EmployeeProfile, user=user)
    admin_data = getattr(profile, "admin_data", None)

    return render(request, "employees/admin_data_detail.html", {
        "employee": profile,
        "admin_data": admin_data,
    })


# --------------------------------------------------
# ADMIN CONTROLLED DATA EDIT (Admin only)
# --------------------------------------------------
@user_passes_test(is_admin)
def admin_data_edit(request, user_id):
    user = get_object_or_404(User, id=user_id)
    profile = get_object_or_404(EmployeeProfile, user=user)

    admin_data, _ = AdminControlledData.objects.get_or_create(profile=profile)

    if request.method == "POST":
        form = EmployeeAdminForm(request.POST, instance=admin_data)
        if form.is_valid():
            admin_data = form.save(commit=False)
            admin_data.profile = profile
            admin_data.save()
            messages.success(request, "Admin-controlled data updated successfully.")
            return redirect("admin_data_detail", user_id=user_id)
    else:
        form = EmployeeAdminForm(instance=admin_data)

    return render(request, "employees/admin_data_edit.html", {
        "employee": profile,
        "form": form,
    })


from django.http import HttpResponse
from reportlab.lib.pagesizes import landscape
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas as pdf_canvas

@login_required
def download_id_card(request, user_id):
    user    = get_object_or_404(User, id=user_id)
    profile = get_object_or_404(EmployeeProfile, user=user)

    # ── SAFE ADMIN DATA ACCESS ──
    try:
        admin_data  = profile.admin_data
        joining     = str(admin_data.joining_date) if admin_data.joining_date else '—'
        role        = str(admin_data.role).capitalize() if admin_data.role else '—'
        designation = str(admin_data.designation).upper() if admin_data.designation else '—'
    except Exception:
        joining     = '—'
        role        = '—'
        designation = '—'

    # ── FULL NAME & INITIAL ──
    full_name = f"{user.first_name} {user.last_name}".strip().title() or user.username.title()
    initial   = (user.first_name[0] if user.first_name else user.username[0]).upper()

    # ── INFO ROWS ──
    info_lines = [
        ("ID",    str(profile.emp_id)),
        ("Email", str(user.email)),
        ("Phone", str(profile.phone or '—')),
        ("Dept",  str(profile.department or '—').upper()),
    ]

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="IDCard_{profile.emp_id}.pdf"'

    width, height = 85.6 * mm, 54 * mm
    c = pdf_canvas.Canvas(response, pagesize=(width, height))

    # ══════════════════════════════════════
    # HEADER
    # ══════════════════════════════════════
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.rect(0, height - 18*mm, width, 18*mm, fill=1, stroke=0)

    c.setFillColor(colors.white)
    c.setFont("Helvetica-Bold", 9)
    c.drawCentredString(width / 2, height - 8*mm, "MANTECH PUBLICATIONS")

    c.setFillColor(colors.HexColor('#93c5fd'))
    c.setFont("Helvetica", 6)
    c.drawCentredString(width / 2, height - 13*mm, "EMPLOYEE IDENTITY CARD")

    # ══════════════════════════════════════
    # FOOTER
    # ══════════════════════════════════════
    c.setFillColor(colors.HexColor('#2563eb'))
    c.rect(0, 0, width, 10*mm, fill=1, stroke=0)  # starts at 0, height 12mm

    c.setFillColor(colors.white)
    c.setFont("Helvetica", 5.5)
    c.drawString(3*mm, 4*mm, f"Joining: {joining}")
    c.drawRightString(width - 3*mm, 4*mm, f"Role: {role}")

    # ══════════════════════════════════════
    # AVATAR — profile image or initials
    # ══════════════════════════════════════
    avatar_x    = 3*mm
    avatar_y    = height - 36*mm
    avatar_size = 16*mm
    cx = avatar_x + avatar_size / 2
    cy = avatar_y + avatar_size / 2
    r  = avatar_size / 2

    if profile.profile_picture:
        try:
            img_path = profile.profile_picture.path

            # Save state before clipping
            c.saveState()

            # Circular clip path
            p = c.beginPath()
            p.circle(cx, cy, r)
            c.clipPath(p, stroke=0, fill=0)

            # Draw image inside clip
            c.drawImage(
                img_path,
                avatar_x,
                avatar_y,
                width=avatar_size,
                height=avatar_size,
                preserveAspectRatio=True,
                mask='auto'
            )

            # Restore state to remove clip
            c.restoreState()

            # Blue border ring over image
            c.setStrokeColor(colors.HexColor('#2563eb'))
            c.setLineWidth(0.5)
            c.circle(cx, cy, r, fill=0, stroke=1)

        except Exception:
            # Fallback to initials if image fails to load
            c.setFillColor(colors.HexColor('#dbeafe'))
            c.circle(cx, cy, r, fill=1, stroke=0)
            c.setStrokeColor(colors.HexColor('#2563eb'))
            c.setLineWidth(0.5)
            c.circle(cx, cy, r, fill=0, stroke=1)
            c.setFillColor(colors.HexColor('#1e3a5f'))
            c.setFont("Helvetica-Bold", 14)
            c.drawCentredString(cx, cy - 2*mm, initial)
    else:
        # No profile picture — draw initials circle
        c.setFillColor(colors.HexColor('#dbeafe'))
        c.circle(cx, cy, r, fill=1, stroke=0)
        c.setStrokeColor(colors.HexColor('#2563eb'))
        c.setLineWidth(0.5)
        c.circle(cx, cy, r, fill=0, stroke=1)
        c.setFillColor(colors.HexColor('#1e3a5f'))
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(cx, cy - 2*mm, initial)

    # ══════════════════════════════════════
    # EMPLOYEE NAME
    # ══════════════════════════════════════
    c.setFillColor(colors.HexColor('#1e3a5f'))
    c.setFont("Helvetica-Bold", 10)
    c.drawString(22*mm, height - 24*mm, full_name)

    # ══════════════════════════════════════
    # DESIGNATION
    # ══════════════════════════════════════
    c.setFillColor(colors.HexColor('#2563eb'))
    c.setFont("Helvetica-Bold", 6.5)
    c.drawString(22*mm, height - 29*mm, designation)

    # ══════════════════════════════════════
    # INFO ROWS
    # ══════════════════════════════════════
    y = height - 32*mm
    for label, value in info_lines:
        c.setFillColor(colors.HexColor('#2563eb'))
        c.setFont("Helvetica-Bold", 6)
        c.drawString(22*mm, y, f"{label}:")
        c.setFillColor(colors.HexColor('#475569'))
        c.setFont("Helvetica", 6)
        c.drawString(34*mm, y, value)
        y -= 3.5*mm

    c.save()
    return response


from reportlab.lib.utils import ImageReader


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────


# ─────────────────────────────────────────────────────────────
# PALETTE
# ─────────────────────────────────────────────────────────────
C_NAVY  = colors.HexColor('#1e3a5f')
C_BLUE  = colors.HexColor('#2563eb')
C_SKY   = colors.HexColor('#93c5fd')
C_PALE  = colors.HexColor('#dbeafe')
C_SLATE = colors.HexColor('#475569')
C_BODY  = colors.HexColor('#334155')
C_WHITE = colors.white
C_GOLD  = colors.HexColor('#f59e0b')
C_LGREY = colors.HexColor('#f1f5f9')


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────

def _header(c, W, H):
    c.setFillColor(C_NAVY)
    c.rect(0, H - 32*mm, W, 32*mm, fill=1, stroke=0)

    c.setFillColor(C_GOLD)
    c.rect(0, H - 1.2*mm, W, 1.2*mm, fill=1, stroke=0)

    p = c.beginPath()
    p.moveTo(W * 0.55, H - 32*mm)
    p.lineTo(W,        H - 32*mm)
    p.lineTo(W,        H - 5*mm)
    p.close()
    c.setFillColor(C_BLUE)
    c.drawPath(p, fill=1, stroke=0)

    c.setStrokeColor(colors.HexColor('#2563eb'))
    c.setLineWidth(0.5)
    c.circle(W - 18*mm, H - 8*mm, 14*mm, fill=0, stroke=1)
    c.circle(W - 18*mm, H - 8*mm, 9*mm,  fill=0, stroke=1)

    c.setFillColor(C_WHITE)
    c.setFont("Helvetica-Bold", 17)
    c.drawString(10*mm, H - 13*mm, "MANTECH PUBLICATIONS")

    c.setFillColor(C_SKY)
    c.setFont("Helvetica", 7)
    c.drawString(10*mm, H - 18.5*mm,
                 "New Delhi – 110001, India   ·   hr.mantechpublications@gmail.com   ·   www.mantechpublications.com")

    c.setStrokeColor(C_BLUE)
    c.setLineWidth(0.6)
    c.line(10*mm, H - 21*mm, W - 10*mm, H - 21*mm)

    badge_w = 52*mm
    badge_x = 10*mm
    badge_y = H - 29.5*mm
    c.setFillColor(C_BLUE)
    c.roundRect(badge_x, badge_y, badge_w, 6.5*mm, 3*mm, fill=1, stroke=0)
    c.setFillColor(C_WHITE)
    c.setFont("Helvetica-Bold", 8)
    c.drawCentredString(badge_x + badge_w / 2, badge_y + 2.2*mm, "EXPERIENCE  LETTER")


def _footer(c, W, ref_no, issue_date):
    c.setFillColor(C_NAVY)
    c.rect(0, 0, W, 12*mm, fill=1, stroke=0)

    c.setFillColor(C_GOLD)
    c.rect(0, 12*mm, W, 0.8*mm, fill=1, stroke=0)

    c.setFillColor(C_WHITE)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawString(10*mm, 7.5*mm, f"Ref. No.:  {ref_no}")
    c.drawRightString(W - 10*mm, 7.5*mm, f"Issue Date:  {issue_date}")

    c.setFillColor(C_SKY)
    c.setFont("Helvetica", 5.8)
    c.drawCentredString(W / 2, 3.5*mm,
                        "This letter is issued on request and is valid only for official purposes.  "
                        "Verification: hr.mantechpublications@gmail.com")


def _section_heading(c, x, y, text, W_content):
    BAR_H = 6*mm
    c.setFillColor(C_PALE)
    c.roundRect(x, y, W_content, BAR_H, 2*mm, fill=1, stroke=0)

    c.setFillColor(C_BLUE)
    c.roundRect(x, y, 3.5*mm, BAR_H, 1.5*mm, fill=1, stroke=0)
    c.rect(x + 1.8*mm, y, 1.7*mm, BAR_H, fill=1, stroke=0)

    c.setFillColor(C_NAVY)
    c.setFont("Helvetica-Bold", 8)
    c.drawString(x + 6*mm, y + 2*mm, text.upper())

    return y - 4*mm


def _detail_cell(c, x, y, label, value, cell_w):
    c.setFillColor(C_BLUE)
    c.setFont("Helvetica-Bold", 6.2)
    c.drawString(x, y, label.upper())

    c.setFillColor(C_NAVY)
    c.setFont("Helvetica-Bold", 8.5)
    c.drawString(x, y - 4.2*mm, value)

    c.setStrokeColor(C_PALE)
    c.setLineWidth(0.5)
    c.line(x, y - 5.5*mm, x + cell_w - 4*mm, y - 5.5*mm)


def _justified_paragraph(c, text, x, y, max_w, size=8.2, leading=5.2*mm):
    words = text.split()
    lines = []
    line  = []

    for word in words:
        trial = " ".join(line + [word])
        if c.stringWidth(trial, "Helvetica", size) <= max_w:
            line.append(word)
        else:
            lines.append(line)
            line = [word]
    if line:
        lines.append(line)

    for i, words_in_line in enumerate(lines):
        if i < len(lines) - 1 and len(words_in_line) > 1:
            total_word_w = sum(c.stringWidth(w, "Helvetica", size) for w in words_in_line)
            gap = (max_w - total_word_w) / (len(words_in_line) - 1)
            cx  = x
            for word in words_in_line:
                c.setFont("Helvetica", size)
                c.setFillColor(C_BODY)
                c.drawString(cx, y, word)
                cx += c.stringWidth(word, "Helvetica", size) + gap
        else:
            c.setFont("Helvetica", size)
            c.setFillColor(C_BODY)
            c.drawString(x, y, " ".join(words_in_line))
        y -= leading

    return y


# ─────────────────────────────────────────────────────────────
# VIEW
# ─────────────────────────────────────────────────────────────

@login_required
def download_experience_letter(request, user_id):
    user    = get_object_or_404(User, id=user_id)
    profile = get_object_or_404(EmployeeProfile, user=user)

    # ── Data ──────────────────────────────────────────────────
    full_name  = f"{user.first_name} {user.last_name}".strip().title() or user.username.title()
    first_name = (user.first_name or full_name.split()[0]).title()

    try:
        admin_data  = profile.admin_data
        joining     = admin_data.joining_date.strftime("%d %B %Y") if admin_data.joining_date else "—"
        designation = str(admin_data.designation).upper()           if admin_data.designation else "—"
        role        = str(admin_data.role).capitalize()             if admin_data.role        else "—"
    except Exception:
        joining     = "—"
        designation = "—"
        role        = "—"

    department     = str(profile.department or "—").upper()
    emp_id         = str(profile.emp_id)
    today          = date.today()
    issue_date     = today.strftime("%d %B %Y")
    relieving_date = issue_date   # replace with actual field if stored
    ref_no         = f"MNTCH/EXP/{today.year}/{emp_id}"

    # ── Canvas setup ──────────────────────────────────────────
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="ExperienceLetter_{emp_id}.pdf"'

    W, H = 210*mm, 297*mm
    c    = pdf_canvas.Canvas(response, pagesize=(W, H))
    MX   = 12*mm
    CW   = W - 2*MX

    # ── HEADER ────────────────────────────────────────────────
    _header(c, W, H)

    # ── REF / DATE strip ──────────────────────────────────────
    strip_y = H - 40*mm
    c.setFillColor(C_LGREY)
    c.rect(0, strip_y, W, 7*mm, fill=1, stroke=0)
    c.setFillColor(C_SLATE)
    c.setFont("Helvetica", 7)
    c.drawString(MX, strip_y + 2.3*mm, f"Ref. No.:  {ref_no}")
    c.drawRightString(W - MX, strip_y + 2.3*mm, f"Date:  {issue_date}")

    y = strip_y - 12*mm

    # ── EMPLOYEE DETAILS (2-column grid) ──────────────────────
    y = _section_heading(c, MX, y, "Employee Details", CW)
    y -= 2*mm

    col_w   = CW / 2
    left_x  = MX
    right_x = MX + col_w

    left_col = [
        ("Employee Name",    full_name),
        ("Employee ID",      emp_id),
        ("Designation",      designation),
    ]
    right_col = [
        ("Department",       department),
        ("Date of Joining",  joining),
        ("Last Working Day", relieving_date),
    ]

    for (ll, lv), (rl, rv) in zip(left_col, right_col):
        _detail_cell(c, left_x,  y, ll, lv, col_w)
        _detail_cell(c, right_x, y, rl, rv, col_w)
        y -= 10.5*mm

    y -= 4*mm
    c.setStrokeColor(C_GOLD)
    c.setLineWidth(0.7)
    c.line(MX, y, W - MX, y)
    y -= 6*mm

    # ── SALUTATION ────────────────────────────────────────────
    c.setFillColor(C_NAVY)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(MX, y, "To Whomsoever It May Concern,")
    y -= 7*mm

    # ── BODY ──────────────────────────────────────────────────
    y = _section_heading(c, MX, y, "Certificate of Employment", CW)
    y -= 3*mm

    paras = [
        (
            f"This is to certify that {full_name} (Employee ID: {emp_id}) was employed with "
            f"Mantech Publications, New Delhi, as {designation} in the {department} Department "
            f"from {joining} to {relieving_date}."
        ),
        (
            f"During the tenure at Mantech Publications, {first_name} demonstrated exemplary "
            f"professional conduct, strong subject-matter expertise, and a consistent commitment "
            f"to delivering high-quality work. {first_name} was a valued member of the {department} "
            f"team and contributed meaningfully to editorial projects and organisational goals."
        ),
        (
            f"{first_name} has been duly relieved from all duties and responsibilities with effect "
            f"from {relieving_date}. We confirm that there are no outstanding dues or liabilities "
            f"against {full_name} as on the date of this letter."
        ),
        (
            f"We wish {first_name} the very best in all future endeavours and wholeheartedly "
            f"recommend {full_name} to any prospective employer."
        ),
    ]

    for para in paras:
        y = _justified_paragraph(c, para, MX, y, CW)
        y -= 4*mm

    y -= 3*mm

    # ── SIGNATURE BLOCK ───────────────────────────────────────
    sig_h = 28*mm
    sig_y = y - sig_h

    c.setFillColor(C_LGREY)
    c.roundRect(MX, sig_y, CW, sig_h, 2*mm, fill=1, stroke=0)

    c.setFillColor(C_BLUE)
    c.roundRect(MX, sig_y, 3*mm, sig_h, 1.5*mm, fill=1, stroke=0)
    c.rect(MX + 1.5*mm, sig_y, 1.5*mm, sig_h, fill=1, stroke=0)

    c.setFillColor(C_BLUE)
    c.setFont("Helvetica-Bold", 6.5)
    c.drawString(MX + 7*mm, sig_y + sig_h - 6*mm, "AUTHORISED SIGNATORY")

    sig_line_x   = MX + 7*mm
    sig_line_y   = sig_y + 12*mm
    sig_line_end = MX + 68*mm
    c.setStrokeColor(C_NAVY)
    c.setLineWidth(0.7)
    c.line(sig_line_x, sig_line_y, sig_line_end, sig_line_y)

    c.setFillColor(C_NAVY)
    c.setFont("Helvetica-Bold", 9)
    c.drawString(MX + 7*mm, sig_line_y - 5*mm, "Ms. Shivani")

    c.setFillColor(C_BLUE)
    c.setFont("Helvetica", 7.5)
    c.drawString(MX + 7*mm, sig_line_y - 9.5*mm, "Head – Human Resources")

    c.setFillColor(C_SLATE)
    c.setFont("Helvetica", 7)
    c.drawString(MX + 7*mm, sig_line_y - 13.5*mm, "Mantech Publications")

    # Seal circle (right side of sig card)
    stamp_cx = W - MX - 20*mm
    stamp_cy = sig_y + sig_h / 2
    c.setStrokeColor(C_NAVY)
    c.setLineWidth(0.8)
    c.setFillColor(colors.HexColor('#f0f6ff'))
    c.circle(stamp_cx, stamp_cy, 14*mm, fill=1, stroke=1)
    c.setStrokeColor(C_BLUE)
    c.setLineWidth(0.4)
    c.circle(stamp_cx, stamp_cy, 11.5*mm, fill=0, stroke=1)
    c.setFillColor(C_NAVY)
    c.setFont("Helvetica-Bold", 5.5)
    c.drawCentredString(stamp_cx, stamp_cy + 1.5*mm, "OFFICIAL")
    c.drawCentredString(stamp_cx, stamp_cy - 2.5*mm, "SEAL")

    # ── FOOTER ────────────────────────────────────────────────
    _footer(c, W, ref_no, issue_date)

    c.save()
    return response


import os
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST


# ── Constants ──────────────────────────────────────────────────────────────────

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif", "webp"}
MAX_FILE_SIZE_MB = 5
MAX_FILE_SIZE_BYTES = MAX_FILE_SIZE_MB * 1024 * 1024


# ── Helpers ────────────────────────────────────────────────────────────────────

def _allowed_file(filename: str) -> bool:
    """Return True if the file extension is in the allowed set."""
    return (
        "." in filename
        and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS
    )


# ── View ───────────────────────────────────────────────────────────────────────

@login_required
@require_POST
def update_profile_picture(request):
    """
    Handle two sub-actions coming from the profile-picture modal:

        update_type == "upload"  →  validate & save an uploaded image file
        update_type == "avatar"  →  save the chosen avatar name to the profile
    """
    update_type = request.POST.get("update_type", "").strip()
    redirect_url = request.POST.get("next") or None   # fallback redirect

    # ── Branch: file upload ────────────────────────────────────────────────────
    if update_type == "upload":
        image_file = request.FILES.get("profile_picture")

        if not image_file:
            messages.error(request, "No file was selected. Please choose an image.")
            return redirect("employee_detail", user_id=request.user.id)


        # Extension check
        if not _allowed_file(image_file.name):
            messages.error(
                request,
                f"Unsupported file type. Allowed types: {', '.join(ALLOWED_EXTENSIONS).upper()}.",
            )
            return redirect("employee_detail", user_id=request.user.id)


        # Size check
        if image_file.size > MAX_FILE_SIZE_BYTES:
            messages.error(
                request,
                f"File is too large. Maximum allowed size is {MAX_FILE_SIZE_MB} MB.",
            )
            return redirect("employee_detail", user_id=request.user.id)


        # Delete old uploaded file from disk before replacing
        profile = request.user.employee_profile        
        if profile.profile_picture:
            try:
                old_path = profile.profile_picture.path
                if os.path.isfile(old_path):
                    os.remove(old_path)
            except Exception:
                pass  # storage backend may not expose .path (e.g. S3)
 
        profile.profile_picture = image_file
 
        # Clear avatar selection if you add avatar_name field later
        if hasattr(profile, "avatar_name"):
            profile.avatar_name = ""
 
        profile.save()

        messages.success(request, "Your profile picture has been updated successfully.")
        return redirect("employee_detail", user_id=request.user.id)

    # ── Branch: avatar selection ───────────────────────────────────────────────
    elif update_type == "avatar":
        import os

        avatar_name = request.POST.get("avatar_name", "").strip()
        avatar_name = os.path.basename(avatar_name)
        profile = request.user.employee_profile
        
        if not avatar_name:
            messages.error(request, "No avatar was selected. Please click one and try again.")
            return redirect("employee_detail", user_id=request.user.id)
 
        # ── IMPORTANT ──────────────────────────────────────────────────────────
        # Your Profile model has no avatar_name field yet.
        # You have two options:
        #
        # OPTION A (recommended): Add avatar_name to your Profile model
        #   avatar_name = models.CharField(max_length=100, null=True, blank=True)
        #   Then run: python manage.py makemigrations && python manage.py migrate
        #
        # OPTION B: Copy the avatar file into employee_photos/ and save it as
        #   profile_picture so no model change is needed (handled below).
        # ── OPTION B logic (no migration needed) ───────────────────────────────
        from django.conf import settings
        import shutil
        from django.core.files import File
 
        avatar_source = os.path.join(
            settings.STATICFILES_DIRS[0], "icons", avatar_name
        )  # adjust path to where your avatar icons live
 
        if not os.path.isfile(avatar_source):
            # Try BASE_DIR/static/icons/ as fallback
            avatar_source = os.path.join(settings.BASE_DIR, "static", "icons", avatar_name)
 
        if not os.path.isfile(avatar_source):
            messages.error(request, "Avatar file not found on the server.")
            return redirect("employee_detail", user_id=request.user.id)
 
        # Delete old uploaded file
        if profile.profile_picture:
            try:
                old_path = profile.profile_picture.path
                if os.path.isfile(old_path):
                    os.remove(old_path)
            except Exception:
                pass
 
        # Save avatar image into profile_picture (copies into employee_photos/)
        with open(avatar_source, "rb") as f:
            profile.profile_picture.save(avatar_name, File(f), save=True)
 
        messages.success(request, "Your avatar has been updated successfully.")
        return redirect("employee_detail", user_id=request.user.id)

    # ── Fallback: unknown update_type ─────────────────────────────────────────
    messages.error(request, "Invalid request. Please try again.")
    return redirect("employee_detail", user_id=request.user.id)